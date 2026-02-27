"""
Excel 리포트 생성 (v5.0.0 - 3시트 압축)
설계서 섹션 8 기준

v5.0.0 변경:
- 7시트 -> 3시트로 압축 (1_요약, 2_소재TIER, 3_지점별)
- 이모지 제거, 깔끔한 스타일
- 배경색 제거 (헤더만 연한 회색)
- 숫자 포맷 통일
"""
import pandas as pd
import numpy as np
import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[WARNING] openpyxl not installed -> pip install openpyxl")

# v4.0.0: 차트는 별도 HTML 파일로 분리 - Excel에는 차트 미삽입
# 차트 함수는 build_html_charts.py에서 Plotly로 생성
CHARTS_AVAILABLE = False
print("[Excel] v4.0.0 - Data only mode (charts in separate HTML file)")


# 스타일 정의 (v5.0.0 - 심플)
HEADER_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
HEADER_FONT = Font(bold=True, color="333333")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# v5.0.0: 심플 스타일 - KPI 카드
KPI_CARD_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
KPI_VALUE_FONT = Font(bold=True, size=18, color='333333')
KPI_LABEL_FONT = Font(bold=False, size=10, color='666666')

# 섹션 헤더 스타일
SECTION_HEADER_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
SECTION_HEADER_FONT = Font(bold=True, size=12, color="333333")


def clean_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """
    DataFrame에서 NaN/inf 값을 Excel 호환 값으로 변환
    """
    df = df.copy()
    for col in df.columns:
        if df[col].dtype in ['float64', 'float32']:
            # inf를 NaN으로 변환 후 빈 문자열로
            df[col] = df[col].replace([np.inf, -np.inf], np.nan)
            df[col] = df[col].fillna('')
        elif df[col].dtype == 'object':
            df[col] = df[col].fillna('')
    return df


def style_header(ws, row_num=1):
    """헤더 행 스타일링"""
    for cell in ws[row_num]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    # 헤더 행 높이 설정
    ws.row_dimensions[row_num].height = 25


def auto_column_width(ws, min_width=10, max_width=50):
    """컬럼 너비 자동 조정 (개선)"""
    # 컬럼별 최적 너비 계산
    COLUMN_WIDTH_MAP = {
        'TIER': 10, '소재명': 30, '소재유형': 12, '소재구분': 10,
        'CPA': 12, 'CTR': 8, 'CVR': 8, '랜딩률': 8,
        '총비용': 15, '총전환': 10, '총클릭': 10, '집행일수': 10,
        '지점': 10, '나이대': 10, '날짜': 12, 'date': 12,
    }

    for column_cells in ws.columns:
        col_letter = get_column_letter(column_cells[0].column)
        header_value = str(column_cells[0].value or '')

        # 미리 정의된 너비가 있으면 사용
        if header_value in COLUMN_WIDTH_MAP:
            ws.column_dimensions[col_letter].width = COLUMN_WIDTH_MAP[header_value]
        else:
            # 내용 기반 자동 계산
            length = max(len(str(cell.value or '')) for cell in column_cells)
            adjusted_width = min(max(length + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted_width


def set_row_heights(ws, header_height=25, data_height=20, start_row=2):
    """행 높이 일괄 설정 (가독성 향상)"""
    ws.row_dimensions[1].height = header_height
    for row_num in range(start_row, ws.max_row + 1):
        ws.row_dimensions[row_num].height = data_height


def apply_number_alignment(ws, start_row=2):
    """숫자 우측 정렬, 텍스트 좌측 정렬 자동 적용"""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif isinstance(cell.value, str):
                cell.alignment = Alignment(horizontal='left', vertical='center')


def apply_tier_style(ws, tier_col_idx, start_row=2):
    """v5.0.0: TIER 컬럼 볼드 처리만 (색상 없음)"""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        tier_cell = row[tier_col_idx - 1]
        if tier_cell.value:
            tier_cell.font = Font(bold=True)
            tier_cell.alignment = Alignment(horizontal='center', vertical='center')


def format_change_rate(value):
    """변화율 포맷팅 (+/- 기호)"""
    if pd.isna(value) or value == '' or value is None:
        return ''
    try:
        val = float(value)
        if val > 0:
            return f"+{val:.1f}%"
        elif val < 0:
            return f"{val:.1f}%"
        else:
            return "0.0%"
    except (ValueError, TypeError):
        return str(value)


def df_to_sheet(ws, df, apply_border=True, apply_alignment=True, set_heights=True):
    """DataFrame을 시트에 작성 (가독성 개선)"""
    # NaN/inf 값 정리
    df = clean_dataframe_for_excel(df)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if apply_border:
                cell.border = THIN_BORDER
            # 숫자 포맷팅
            if isinstance(value, float):
                col_name = str(df.columns[c_idx-1]) if c_idx <= len(df.columns) else ''
                if 'CPA' in col_name or '비용' in col_name or '총비용' in col_name:
                    cell.number_format = '#,##0'
                elif 'CTR' in col_name or 'CVR' in col_name or '률' in col_name or '비중' in col_name:
                    # CTR/CVR은 이미 퍼센트 값 (0.76 = 0.76%)이므로 리터럴 % 사용
                    cell.number_format = '0.00"%"'
                elif '점수' in col_name:
                    cell.number_format = '0.00'
                elif '전환' in col_name or '클릭' in col_name or '노출' in col_name:
                    cell.number_format = '#,##0'
            elif isinstance(value, int):
                cell.number_format = '#,##0'

    style_header(ws)
    auto_column_width(ws)

    # 가독성 개선 옵션
    if apply_alignment:
        apply_number_alignment(ws, start_row=2)
    if set_heights:
        set_row_heights(ws, header_height=25, data_height=20)


def create_summary_sheet(ws, creative_df, age_df, df_valid, off_df, hook_data=None):
    """v5.0.0: 요약 대시보드 시트 생성 (이모지 제거, 심플 스타일)"""
    ws.title = "1_요약"

    # 기본 정보 계산
    total_cost = df_valid['cost'].sum()
    total_conversions = df_valid['conversions'].sum()
    total_clicks = df_valid['clicks'].sum()
    total_impressions = df_valid['impressions'].sum()
    avg_cpa = total_cost / total_conversions if total_conversions > 0 else 0
    avg_ctr = (total_clicks / total_impressions * 100) if total_impressions > 0 else 0
    avg_cvr = (total_conversions / total_clicks * 100) if total_clicks > 0 else 0

    # 분석 기간
    date_min = df_valid['date'].min()
    date_max = df_valid['date'].max()

    # === 1행: 리포트 제목 ===
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "TikTok 광고 분석 리포트"
    title_cell.font = Font(bold=True, size=18, color='333333')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # === 2행: 분석 기간 및 생성일 ===
    ws.merge_cells('A2:D2')
    ws['A2'].value = f"분석 기간: {date_min.strftime('%Y-%m-%d')} ~ {date_max.strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(size=11, color='666666')

    ws.merge_cells('E2:H2')
    ws['E2'].value = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['E2'].font = Font(size=11, color='666666')
    ws['E2'].alignment = Alignment(horizontal='right')
    ws.row_dimensions[2].height = 25

    # === 3행: 빈 행 (여백) ===
    ws.row_dimensions[3].height = 10

    # === 4-5행: KPI 카드 (4개) - 라벨 + 값 별도 행 ===
    kpi_data = [
        ('A', '총 광고비', f"{total_cost:,.0f}원"),
        ('C', '총 전환수', f"{total_conversions:,.0f}건"),
        ('E', '평균 CPA', f"{avg_cpa:,.0f}원"),
        ('G', '평균 CTR', f"{avg_ctr:.2f}%"),
    ]

    for col, label, value in kpi_data:
        # 라벨 행 (4행) - 2열 병합
        next_col = chr(ord(col) + 1)  # A->B, C->D 등
        ws.merge_cells(f'{col}4:{next_col}4')
        ws[f'{col}4'].value = label
        ws[f'{col}4'].font = KPI_LABEL_FONT
        ws[f'{col}4'].fill = KPI_CARD_FILL
        ws[f'{col}4'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}4'].border = THIN_BORDER

        # 값 행 (5행) - 2열 병합
        ws.merge_cells(f'{col}5:{next_col}5')
        ws[f'{col}5'].value = value
        ws[f'{col}5'].font = KPI_VALUE_FONT
        ws[f'{col}5'].fill = KPI_CARD_FILL
        ws[f'{col}5'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}5'].border = THIN_BORDER

    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 35

    # === 6행: 빈 행 (여백) ===
    ws.row_dimensions[6].height = 15

    # === 7행: 섹션 헤더 - TIER 분포 ===
    ws.merge_cells('A7:D7')
    ws['A7'].value = "TIER 분포"
    ws['A7'].fill = SECTION_HEADER_FILL
    ws['A7'].font = SECTION_HEADER_FONT
    ws['A7'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[7].height = 25

    # === 8행~: TIER 분포 테이블 ===
    tier_order = ['TIER1', 'TIER2', 'TIER3', 'TIER4', 'LOW_VOLUME', 'UNCLASSIFIED']
    tier_dist = creative_df['TIER'].value_counts()

    row_num = 8
    for tier in tier_order:
        count = tier_dist.get(tier, 0)
        if count > 0:
            ws[f'A{row_num}'].value = tier
            ws[f'A{row_num}'].font = Font(bold=True)
            ws[f'A{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{row_num}'].border = THIN_BORDER

            ws[f'B{row_num}'].value = f"{count}개"
            ws[f'B{row_num}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'B{row_num}'].border = THIN_BORDER

            pct = count / len(creative_df) * 100
            ws[f'C{row_num}'].value = f"({pct:.1f}%)"
            ws[f'C{row_num}'].font = Font(color='666666')
            ws[f'C{row_num}'].border = THIN_BORDER

            ws.row_dimensions[row_num].height = 22
            row_num += 1

    # === 데이터 품질 참고사항 섹션 ===
    row_num += 1
    ws.merge_cells(f'A{row_num}:D{row_num}')
    ws[f'A{row_num}'].value = "데이터 품질 참고사항"
    ws[f'A{row_num}'].fill = SECTION_HEADER_FILL
    ws[f'A{row_num}'].font = SECTION_HEADER_FONT
    for cell in [f'B{row_num}', f'C{row_num}', f'D{row_num}']:
        ws[cell].fill = SECTION_HEADER_FILL
    ws.row_dimensions[row_num].height = 25
    row_num += 1

    # 참고사항 데이터
    ref_data = [
        ("분석 소재", f"{len(creative_df)}개"),
        ("OFF 소재", f"{len(off_df) if off_df is not None and len(off_df) > 0 else 0}개"),
    ]

    if 'attribution_caution' in df_valid.columns:
        caution_count = df_valid['attribution_caution'].sum()
        if caution_count > 0:
            ref_data.append(("귀속 주의", f"{caution_count}건 (클릭=0, 전환>0)"))

    for label, value in ref_data:
        ws[f'A{row_num}'].value = label
        ws[f'A{row_num}'].font = Font(color='666666')
        ws[f'B{row_num}'].value = value
        ws[f'B{row_num}'].alignment = Alignment(horizontal='left')
        ws.row_dimensions[row_num].height = 20
        row_num += 1

    # 컬럼 너비 설정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15


def create_tier_sheet(ws, creative_df):
    """v5.0.0: 소재 TIER 분석 시트 생성 (심플 컬럼)"""
    ws.title = "2_소재TIER"

    # v5.0.0: 필수 컬럼만 출력 (불필요 컬럼 제거)
    columns = ['소재명', 'TIER', 'CPA', 'CTR', 'CVR', '랜딩률', '총비용', '총전환', '총클릭', '집행일수']

    # 존재하는 컬럼만 선택
    available_cols = [c for c in columns if c in creative_df.columns]
    df_output = creative_df[available_cols].copy()

    # TIER 순서로 정렬 (성과 좋은 순)
    tier_order = ['TIER1', 'TIER2', 'TIER3', 'TIER4', 'LOW_VOLUME', 'UNCLASSIFIED']
    df_output['TIER_order'] = df_output['TIER'].map({t: i for i, t in enumerate(tier_order)})
    df_output = df_output.sort_values(['TIER_order', 'CPA']).drop(columns=['TIER_order'])

    # 소재명 길이 제한 (가독성)
    if '소재명' in df_output.columns:
        df_output['소재명'] = df_output['소재명'].apply(
            lambda x: x[:25] + '...' if isinstance(x, str) and len(x) > 28 else x
        )

    df_to_sheet(ws, df_output, apply_alignment=True, set_heights=True)

    # TIER 볼드 처리만 (색상 없음)
    if 'TIER' in available_cols:
        tier_col_idx = list(df_output.columns).index('TIER') + 1
        apply_tier_style(ws, tier_col_idx)


def create_branch_sheet(ws, df_valid):
    """v5.0.0: 지점별 성과 시트 생성 (이모지 제거)"""
    ws.title = "3_지점별"

    # 지점별 집계
    branch_summary = df_valid.groupby('지점').agg(
        총비용=('cost', 'sum'),
        총전환=('conversions', 'sum'),
        총클릭=('clicks', 'sum'),
        총노출=('impressions', 'sum'),
    ).reset_index()

    branch_summary['CPA'] = (branch_summary['총비용'] / branch_summary['총전환'].replace(0, np.nan)).round(0)
    branch_summary['CTR'] = (branch_summary['총클릭'] / branch_summary['총노출'].replace(0, np.nan) * 100).round(2)
    branch_summary['CVR'] = (branch_summary['총전환'] / branch_summary['총클릭'].replace(0, np.nan) * 100).round(2)

    # CPA 순으로 정렬 (성과 좋은 순)
    branch_summary = branch_summary.sort_values('CPA', na_position='last').reset_index(drop=True)

    # 순위 및 평가 컬럼 추가 (이모지 제거)
    def get_rank_text(idx):
        return f"{idx + 1}위"

    def get_evaluation(cpa, median_cpa):
        if pd.isna(cpa) or cpa == '':
            return ''
        if cpa <= median_cpa * 0.8:
            return '우수'
        elif cpa <= median_cpa:
            return '양호'
        elif cpa <= median_cpa * 1.2:
            return '보통'
        else:
            return '개선필요'

    # 중앙값 계산 (평가 기준)
    valid_cpa = branch_summary['CPA'].dropna()
    median_cpa = valid_cpa.median() if len(valid_cpa) > 0 else 0

    branch_summary['순위'] = [get_rank_text(i) for i in range(len(branch_summary))]
    branch_summary['평가'] = branch_summary['CPA'].apply(lambda x: get_evaluation(x, median_cpa))

    # 컬럼 순서 재정렬
    col_order = ['지점', '순위', 'CPA', 'CTR', 'CVR', '총비용', '총전환', '평가']
    available_cols = [c for c in col_order if c in branch_summary.columns]
    branch_summary = branch_summary[available_cols]

    # NaN/inf 값 정리
    branch_summary = branch_summary.fillna('')

    df_to_sheet(ws, branch_summary)




def build_excel(output_dir: str, creative_df: pd.DataFrame, age_df: pd.DataFrame,
                df_valid: pd.DataFrame, off_df: pd.DataFrame = None,
                hook_type_df: pd.DataFrame = None, hook_strict_df: pd.DataFrame = None,
                target_cpa_map: dict = None):
    """
    v5.0.0: Excel 리포트 생성 (3개 시트 - 1_요약, 2_소재TIER, 3_지점별)
    이모지 제거, 심플 스타일
    """
    if not OPENPYXL_AVAILABLE:
        print("[ERROR] openpyxl not installed. pip install openpyxl")
        return None

    wb = Workbook()

    # 1. 요약 대시보드
    ws_summary = wb.active
    create_summary_sheet(ws_summary, creative_df, age_df, df_valid, off_df)

    # 2. 소재 TIER 분석
    ws_tier = wb.create_sheet()
    create_tier_sheet(ws_tier, creative_df)

    # 3. 지점별 성과
    ws_branch = wb.create_sheet()
    create_branch_sheet(ws_branch, df_valid)

    # 저장
    os.makedirs(output_dir, exist_ok=True)
    today = datetime.now().strftime('%Y%m%d')
    output_path = os.path.join(output_dir, f"tiktok_analysis_{today}.xlsx")

    # Generate analysis_validation.json
    generate_validation_json(output_dir, df_valid, creative_df, age_df)

    wb.save(output_path)
    print(f"[OK] Excel report generated (v5.0.0 - 3 sheets) -> {output_path}")

    return output_path


def generate_validation_json(output_dir: str, df_valid: pd.DataFrame,
                            creative_df: pd.DataFrame, age_df: pd.DataFrame):
    """
    Generate analysis_validation.json for data integrity check
    """
    import json

    # Calculate totals
    raw_total_conv = df_valid['conversions'].sum()
    raw_total_cost = df_valid['cost'].sum()

    # Creative-level totals
    creative_conv = creative_df['총전환'].sum() if '총전환' in creative_df.columns else 0

    # Daily totals (from df_valid)
    daily_conv = df_valid.groupby('date')['conversions'].sum().sum()

    # Branch totals
    branch_conv = df_valid.groupby('지점')['conversions'].sum().sum() if '지점' in df_valid.columns else 0

    # Check matches
    all_match = (
        abs(raw_total_conv - creative_conv) <= 1 and
        abs(raw_total_conv - daily_conv) <= 1 and
        abs(raw_total_conv - branch_conv) <= 1
    )

    validation = {
        "generated_at": datetime.now().strftime('%Y-%m-%d %H:%M'),
        "validation": {
            "raw_total_conv": int(raw_total_conv),
            "raw_total_cost": float(raw_total_cost),
            "creative_conv_sum": int(creative_conv),
            "daily_conv_sum": int(daily_conv),
            "branch_conv_sum": int(branch_conv),
            "all_match": bool(all_match),  # Convert numpy bool to Python bool
            "analysis_total_conv": int(creative_conv)
        }
    }

    # Save
    os.makedirs(output_dir, exist_ok=True)
    json_path = os.path.join(output_dir, "analysis_validation.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(validation, f, ensure_ascii=False, indent=2)

    print(f"[OK] Validation JSON -> {json_path}")

    if not all_match:
        print(f"[WARNING] Data mismatch detected!")
        print(f"  Raw: {raw_total_conv}, Creative: {creative_conv}, Daily: {daily_conv}, Branch: {branch_conv}")

    return validation


def main(data_dir: str = "output", output_dir: str = None):
    """
    메인 실행 함수
    """
    if output_dir is None:
        today = datetime.now().strftime('%Y%m%d')
        output_dir = os.path.join(data_dir, today)

    # 데이터 로드
    creative_df = pd.read_parquet(os.path.join(data_dir, "creative_tier.parquet"))
    print(f"[Load] Creative TIER: {len(creative_df)}")

    try:
        age_df = pd.read_parquet(os.path.join(data_dir, "age_analysis.parquet"))
        print(f"[Load] Age analysis: {len(age_df)}")
    except FileNotFoundError:
        age_df = pd.DataFrame()

    try:
        off_df = pd.read_parquet(os.path.join(data_dir, "creative_off.parquet"))
        print(f"[Load] OFF creatives: {len(off_df)}")
    except FileNotFoundError:
        off_df = pd.DataFrame()

    try:
        df_valid = pd.read_parquet(os.path.join(data_dir, "parsed.parquet"))
        df_valid = df_valid[df_valid['parse_status'] == 'OK']
        print(f"[Load] Parsed data: {len(df_valid)} rows")
    except FileNotFoundError:
        print("[ERROR] parsed.parquet not found")
        return None

    # 훅 비교 데이터 로드
    try:
        hook_type_df = pd.read_parquet(os.path.join(data_dir, "hook_type_comparison.parquet"))
        print(f"[Load] Hook type comparison: {len(hook_type_df)}")
    except FileNotFoundError:
        hook_type_df = None

    try:
        hook_strict_df = pd.read_parquet(os.path.join(data_dir, "hook_strict_pairs.parquet"))
        print(f"[Load] Hook strict pairs: {len(hook_strict_df)}")
    except FileNotFoundError:
        hook_strict_df = None

    return build_excel(output_dir, creative_df, age_df, df_valid, off_df, hook_type_df, hook_strict_df)


if __name__ == "__main__":
    import sys
    data_directory = sys.argv[1] if len(sys.argv) > 1 else "output"
    output_directory = sys.argv[2] if len(sys.argv) > 2 else None
    main(data_directory, output_directory)

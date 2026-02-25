"""
Excel 리포트 생성
설계서 섹션 8 기준

7개 시트:
1. 요약 대시보드 - 전체 KPI + TIER 분포
2. 소재 TIER 분석 - 소재별 TIER + 전 지표 + 액션
3. 훅 개선 효과 - Strict 쌍 비교 + 소재유형별 집계 비교
4. 지점 컨텍스트 - 지점별 성과 요약
5. 나이대 분석 - 예산 효율 점수
6. 일별 트렌드 - 일별 KPI 추이
7. OFF 소재 - OFF 소재 마지막 성과 보존
"""
import pandas as pd
import numpy as np
import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[WARNING] openpyxl not installed -> pip install openpyxl")

# Chart functions - handle import from same directory
import sys
import os as os_module
_current_dir = os_module.path.dirname(os_module.path.abspath(__file__))
if _current_dir not in sys.path:
    sys.path.insert(0, _current_dir)

try:
    from build_charts import (
        add_tier_donut,
        add_branch_cpa_bar,
        add_creative_bubble_chart,
        add_hook_comparison_chart,
        add_age_efficiency_chart,
        add_age_type_heatmap,
        add_daily_combo_chart,
        add_daily_ctr_line,
        # v3.2.0 신규 차트
        add_type_radar_chart,
        add_fatigue_line_chart,
        add_daily_cpa_trend_with_target,
    )
    CHARTS_AVAILABLE = True
    print("[Charts] Chart functions loaded (v3.2.0)")
except ImportError as e:
    CHARTS_AVAILABLE = False
    print(f"[WARNING] build_charts.py not found - charts will be skipped: {e}")


# 스타일 정의
TIER_COLORS = {
    'TIER1': '90EE90',  # 연한 녹색
    'TIER2': 'ADD8E6',  # 연한 파랑
    'TIER3': 'FFFFE0',  # 연한 노랑
    'TIER4': 'FFB6C1',  # 연한 빨강
    'LOW_VOLUME': 'D3D3D3',  # 회색
    'UNCLASSIFIED': 'E8E8E8',  # 밝은 회색
}

# TIER별 글자 색상 (가독성 향상)
TIER_TEXT_COLORS = {
    'TIER1': '155724',  # 진한 녹색
    'TIER2': '004085',  # 진한 파랑
    'TIER3': '856404',  # 진한 노랑
    'TIER4': '721C24',  # 진한 빨강
    'LOW_VOLUME': '383D41',  # 진한 회색
    'UNCLASSIFIED': '6C757D',  # 회색
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# KPI 카드 스타일 (요약 대시보드용)
KPI_CARD_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
KPI_CARD_BORDER = Border(
    left=Side(style='medium', color='3498DB'),
    right=Side(style='medium', color='3498DB'),
    top=Side(style='medium', color='3498DB'),
    bottom=Side(style='medium', color='3498DB')
)
KPI_VALUE_FONT = Font(bold=True, size=18, color='2C3E50')
KPI_LABEL_FONT = Font(bold=False, size=10, color='7F8C8D')
KPI_UNIT_FONT = Font(bold=False, size=9, color='95A5A6')

# 섹션 헤더 스타일
SECTION_HEADER_FILL = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
SECTION_HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")

# 상태 색상 (변화율 등)
STATUS_COLORS = {
    'positive': PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    'negative': PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
    'warning': PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
    'neutral': PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
}

# 효율 판정 스타일
EFFICIENCY_STYLES = {
    '✅ 우수': {'fill': 'E8F5E9', 'color': '2E7D32'},
    '➡️ 양호': {'fill': 'FFFFFF', 'color': '333333'},
    '⬇️ 낮음': {'fill': 'FFF8E1', 'color': 'F57C00'},
    '🔴 비효율': {'fill': 'FFEBEE', 'color': 'C62828'},
}

# 훅 판정 스타일
HOOK_VERDICT_STYLES = {
    '유효': {'fill': 'E8F5E9', 'color': '2E7D32', 'icon': '✓'},
    '부분효과': {'fill': 'FFF3E0', 'color': 'EF6C00', 'icon': '△'},
    '효과없음': {'fill': 'FFEBEE', 'color': 'C62828', 'icon': '✗'},
}


def clean_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """
    DataFrame에서 NaN/inf 값을 Excel 호환 값으로 변환
    """
    df = df.copy()
    for col in df.columns:
        if df[col].dtype in ['float64', 'float32']:
            # inf를 NaN으로 변환 후 빈 문자열로
            df[col] = df[col].replace([np.inf, -np.inf], np.nan)
            df[col] = df[col].fillna('')
        elif df[col].dtype == 'object':
            df[col] = df[col].fillna('')
    return df


def style_header(ws, row_num=1):
    """헤더 행 스타일링"""
    for cell in ws[row_num]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    # 헤더 행 높이 설정
    ws.row_dimensions[row_num].height = 25


def auto_column_width(ws, min_width=10, max_width=50):
    """컬럼 너비 자동 조정 (개선)"""
    # 컬럼별 최적 너비 계산
    COLUMN_WIDTH_MAP = {
        'TIER': 10, '소재명': 30, '소재유형': 12, '소재구분': 10,
        'CPA': 12, 'CTR': 8, 'CVR': 8, '랜딩률': 8,
        '총비용': 15, '총전환': 10, '총클릭': 10, '집행일수': 10,
        '지점': 10, '나이대': 10, '날짜': 12, 'date': 12,
    }

    for column_cells in ws.columns:
        col_letter = get_column_letter(column_cells[0].column)
        header_value = str(column_cells[0].value or '')

        # 미리 정의된 너비가 있으면 사용
        if header_value in COLUMN_WIDTH_MAP:
            ws.column_dimensions[col_letter].width = COLUMN_WIDTH_MAP[header_value]
        else:
            # 내용 기반 자동 계산
            length = max(len(str(cell.value or '')) for cell in column_cells)
            adjusted_width = min(max(length + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted_width


def set_row_heights(ws, header_height=25, data_height=20, start_row=2):
    """행 높이 일괄 설정 (가독성 향상)"""
    ws.row_dimensions[1].height = header_height
    for row_num in range(start_row, ws.max_row + 1):
        ws.row_dimensions[row_num].height = data_height


def apply_number_alignment(ws, start_row=2):
    """숫자 우측 정렬, 텍스트 좌측 정렬 자동 적용"""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif isinstance(cell.value, str):
                # 상태/판정 텍스트는 중앙 정렬
                if any(icon in str(cell.value) for icon in ['✅', '➡️', '⬇️', '🔴', '✓', '△', '✗', '▲', '▼']):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')


def apply_tier_colors(ws, tier_col_idx, start_row=2):
    """TIER 컬럼에 색상 적용 (개선: TIER 셀만 진하게, 행은 연하게)"""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        tier_cell = row[tier_col_idx - 1]
        tier_value = tier_cell.value
        if tier_value in TIER_COLORS:
            # TIER 셀: 진한 배경 + 굵은 글씨
            tier_fill = PatternFill(start_color=TIER_COLORS[tier_value],
                                    end_color=TIER_COLORS[tier_value],
                                    fill_type="solid")
            tier_cell.fill = tier_fill
            tier_cell.font = Font(bold=True, color=TIER_TEXT_COLORS.get(tier_value, '000000'))
            tier_cell.alignment = Alignment(horizontal='center', vertical='center')

            # 나머지 셀: 연한 배경 (알파 조절 효과)
            light_color = TIER_COLORS[tier_value]
            row_fill = PatternFill(start_color=f"F{light_color[1:]}",
                                   end_color=f"F{light_color[1:]}",
                                   fill_type="solid")
            for cell in row:
                if cell != tier_cell:
                    cell.fill = row_fill


def format_change_rate(value):
    """변화율 포맷팅 (▲/▼ 아이콘 포함)"""
    if pd.isna(value) or value == '' or value is None:
        return ''
    try:
        val = float(value)
        if val > 0:
            return f"▲+{val:.1f}%"
        elif val < 0:
            return f"▼{val:.1f}%"
        else:
            return "0.0%"
    except (ValueError, TypeError):
        return str(value)


def apply_change_rate_colors(ws, col_idx, start_row=2, reverse=False):
    """변화율 셀에 색상 적용 (CPA는 reverse=True: 하락이 좋음)"""
    green_font = Font(color='2E7D32', bold=True)
    red_font = Font(color='C62828', bold=True)

    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        cell = row[col_idx - 1]
        value = str(cell.value) if cell.value else ''
        if '▲' in value:
            cell.font = red_font if reverse else green_font
            cell.fill = STATUS_COLORS['negative'] if reverse else STATUS_COLORS['positive']
        elif '▼' in value:
            cell.font = green_font if reverse else red_font
            cell.fill = STATUS_COLORS['positive'] if reverse else STATUS_COLORS['negative']


def apply_efficiency_styles(ws, col_idx, start_row=2):
    """효율 판정 컬럼에 스타일 적용"""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        cell = row[col_idx - 1]
        value = str(cell.value) if cell.value else ''
        for verdict, style in EFFICIENCY_STYLES.items():
            if verdict in value:
                cell.fill = PatternFill(start_color=style['fill'],
                                        end_color=style['fill'],
                                        fill_type="solid")
                cell.font = Font(bold=True, color=style['color'])
                cell.alignment = Alignment(horizontal='center', vertical='center')
                break


def apply_hook_verdict_styles(ws, col_idx, start_row=2):
    """훅 판정 컬럼에 스타일 적용"""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        cell = row[col_idx - 1]
        value = str(cell.value) if cell.value else ''
        for verdict, style in HOOK_VERDICT_STYLES.items():
            if verdict in value:
                cell.fill = PatternFill(start_color=style['fill'],
                                        end_color=style['fill'],
                                        fill_type="solid")
                cell.font = Font(bold=True, color=style['color'])
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # 아이콘 추가
                if style['icon'] not in value:
                    cell.value = f"{value} {style['icon']}"
                break


def df_to_sheet(ws, df, apply_border=True, apply_alignment=True, set_heights=True):
    """DataFrame을 시트에 작성 (가독성 개선)"""
    # NaN/inf 값 정리
    df = clean_dataframe_for_excel(df)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if apply_border:
                cell.border = THIN_BORDER
            # 숫자 포맷팅
            if isinstance(value, float):
                col_name = str(df.columns[c_idx-1]) if c_idx <= len(df.columns) else ''
                if 'CPA' in col_name or '비용' in col_name or '총비용' in col_name:
                    cell.number_format = '#,##0'
                elif 'CTR' in col_name or 'CVR' in col_name or '률' in col_name or '비중' in col_name:
                    cell.number_format = '0.0%' if value < 1 else '0.00'
                elif '점수' in col_name:
                    cell.number_format = '0.00'
                elif '전환' in col_name or '클릭' in col_name or '노출' in col_name:
                    cell.number_format = '#,##0'
            elif isinstance(value, int):
                cell.number_format = '#,##0'

    style_header(ws)
    auto_column_width(ws)

    # 가독성 개선 옵션
    if apply_alignment:
        apply_number_alignment(ws, start_row=2)
    if set_heights:
        set_row_heights(ws, header_height=25, data_height=20)


def create_summary_sheet(ws, creative_df, age_df, df_valid, off_df, hook_data=None):
    """요약 대시보드 시트 생성 (KPI 카드 형태로 개선)"""
    ws.title = "📊 요약 대시보드"

    # 기본 정보 계산
    total_cost = df_valid['cost'].sum()
    total_conversions = df_valid['conversions'].sum()
    total_clicks = df_valid['clicks'].sum()
    total_impressions = df_valid['impressions'].sum()
    avg_cpa = total_cost / total_conversions if total_conversions > 0 else 0
    avg_ctr = (total_clicks / total_impressions * 100) if total_impressions > 0 else 0
    avg_cvr = (total_conversions / total_clicks * 100) if total_clicks > 0 else 0

    # 분석 기간
    date_min = df_valid['date'].min()
    date_max = df_valid['date'].max()

    # === 1행: 리포트 제목 ===
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "🎯 TikTok 광고 분석 리포트"
    title_cell.font = Font(bold=True, size=18, color='2C3E50')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # === 2행: 분석 기간 및 생성일 ===
    ws.merge_cells('A2:D2')
    ws['A2'].value = f"📅 분석 기간: {date_min.strftime('%Y-%m-%d')} ~ {date_max.strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(size=11, color='7F8C8D')

    ws.merge_cells('E2:H2')
    ws['E2'].value = f"🕐 생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['E2'].font = Font(size=11, color='7F8C8D')
    ws['E2'].alignment = Alignment(horizontal='right')
    ws.row_dimensions[2].height = 25

    # === 3행: 빈 행 (여백) ===
    ws.row_dimensions[3].height = 10

    # === 4-5행: KPI 카드 (4개) - 라벨 + 값 별도 행 ===
    kpi_data = [
        ('A', '총 광고비', f"{total_cost:,.0f}원"),
        ('C', '총 전환수', f"{total_conversions:,.0f}건"),
        ('E', '평균 CPA', f"{avg_cpa:,.0f}원"),
        ('G', '평균 CTR', f"{avg_ctr:.2f}%"),
    ]

    for col, label, value in kpi_data:
        # 라벨 행 (4행) - 2열 병합
        next_col = chr(ord(col) + 1)  # A->B, C->D 등
        ws.merge_cells(f'{col}4:{next_col}4')
        ws[f'{col}4'].value = label
        ws[f'{col}4'].font = KPI_LABEL_FONT
        ws[f'{col}4'].fill = KPI_CARD_FILL
        ws[f'{col}4'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}4'].border = Border(
            left=Side(style='medium', color='3498DB'),
            right=Side(style='medium', color='3498DB'),
            top=Side(style='medium', color='3498DB'),
        )

        # 값 행 (5행) - 2열 병합
        ws.merge_cells(f'{col}5:{next_col}5')
        ws[f'{col}5'].value = value
        ws[f'{col}5'].font = KPI_VALUE_FONT
        ws[f'{col}5'].fill = KPI_CARD_FILL
        ws[f'{col}5'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}5'].border = Border(
            left=Side(style='medium', color='3498DB'),
            right=Side(style='medium', color='3498DB'),
            bottom=Side(style='medium', color='3498DB'),
        )

    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 35

    # === 6행: 빈 행 (여백) ===
    ws.row_dimensions[6].height = 15

    # === 7행: 섹션 헤더 - TIER 분포 ===
    ws.merge_cells('A7:D7')
    ws['A7'].value = "📊 TIER 분포"
    ws['A7'].fill = SECTION_HEADER_FILL
    ws['A7'].font = SECTION_HEADER_FONT
    ws['A7'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[7].height = 25

    # === 8행~: TIER 분포 테이블 ===
    tier_order = ['TIER1', 'TIER2', 'TIER3', 'TIER4', 'LOW_VOLUME', 'UNCLASSIFIED']
    tier_dist = creative_df['TIER'].value_counts()

    row_num = 8
    for tier in tier_order:
        count = tier_dist.get(tier, 0)
        if count > 0:
            ws[f'A{row_num}'].value = tier
            ws[f'A{row_num}'].fill = PatternFill(start_color=TIER_COLORS[tier],
                                                  end_color=TIER_COLORS[tier],
                                                  fill_type="solid")
            ws[f'A{row_num}'].font = Font(bold=True, color=TIER_TEXT_COLORS.get(tier, '000000'))
            ws[f'A{row_num}'].alignment = Alignment(horizontal='center', vertical='center')

            ws[f'B{row_num}'].value = f"{count}개"
            ws[f'B{row_num}'].alignment = Alignment(horizontal='left', vertical='center')

            pct = count / len(creative_df) * 100
            ws[f'C{row_num}'].value = f"({pct:.1f}%)"
            ws[f'C{row_num}'].font = Font(color='7F8C8D')

            ws.row_dimensions[row_num].height = 22
            row_num += 1

    # === 데이터 품질 참고사항 섹션 ===
    row_num += 1
    ws.merge_cells(f'A{row_num}:D{row_num}')
    ws[f'A{row_num}'].value = "⚠️ 데이터 품질 참고사항"
    ws[f'A{row_num}'].fill = SECTION_HEADER_FILL
    ws[f'A{row_num}'].font = SECTION_HEADER_FONT
    for cell in [f'B{row_num}', f'C{row_num}', f'D{row_num}']:
        ws[cell].fill = SECTION_HEADER_FILL
    ws.row_dimensions[row_num].height = 25
    row_num += 1

    # 참고사항 데이터
    ref_data = [
        ("분석 소재", f"{len(creative_df)}개"),
        ("OFF 소재", f"{len(off_df) if off_df is not None and len(off_df) > 0 else 0}개"),
    ]

    if 'attribution_caution' in df_valid.columns:
        caution_count = df_valid['attribution_caution'].sum()
        if caution_count > 0:
            ref_data.append(("⚡ 귀속 주의", f"{caution_count}건 (클릭=0, 전환>0)"))

    for label, value in ref_data:
        ws[f'A{row_num}'].value = label
        ws[f'A{row_num}'].font = Font(color='7F8C8D')
        ws[f'B{row_num}'].value = value
        ws[f'B{row_num}'].alignment = Alignment(horizontal='left')
        ws.row_dimensions[row_num].height = 20
        row_num += 1

    # 컬럼 너비 설정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15


def create_tier_sheet(ws, creative_df):
    """소재 TIER 분석 시트 생성 (가독성 개선)"""
    ws.title = "🎬 소재 TIER 분석"

    # 출력할 컬럼 선택 및 순서 (그룹별 정리)
    # 그룹 A: 핵심 정보
    core_cols = ['TIER', '소재명', '소재유형', 'CPA', 'CTR', 'CVR', '랜딩률']
    # 그룹 B: 상세 정보
    detail_cols = ['총비용', '총전환', '총클릭', '집행일수', '집행지점목록']
    # 그룹 C: 참고 정보
    ref_cols = ['지점편중주석', '지점별_상대평가', 'TIER_근거']

    columns = core_cols + detail_cols + ref_cols

    # 존재하는 컬럼만 선택
    available_cols = [c for c in columns if c in creative_df.columns]
    df_output = creative_df[available_cols].copy()

    # TIER 순서로 정렬 (성과 좋은 순)
    tier_order = ['TIER1', 'TIER2', 'TIER3', 'TIER4', 'LOW_VOLUME', 'UNCLASSIFIED']
    df_output['TIER_order'] = df_output['TIER'].map({t: i for i, t in enumerate(tier_order)})
    df_output = df_output.sort_values(['TIER_order', 'CPA']).drop(columns=['TIER_order'])

    # 집행지점목록을 문자열로 변환
    if '집행지점목록' in df_output.columns:
        df_output['집행지점목록'] = df_output['집행지점목록'].apply(
            lambda x: ', '.join(x) if isinstance(x, list) else str(x)
        )

    # 소재명 길이 제한 (가독성)
    if '소재명' in df_output.columns:
        df_output['소재명'] = df_output['소재명'].apply(
            lambda x: x[:25] + '...' if isinstance(x, str) and len(x) > 28 else x
        )

    df_to_sheet(ws, df_output, apply_alignment=True, set_heights=True)

    # TIER 색상 적용 (개선된 버전)
    tier_col_idx = list(df_output.columns).index('TIER') + 1
    apply_tier_colors(ws, tier_col_idx)

    # 컬럼 그룹별 스타일 구분
    core_count = len([c for c in core_cols if c in available_cols])
    detail_start = core_count + 1

    # 핵심 컬럼 헤더 강조
    for col_idx in range(1, core_count + 1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True, color='FFFFFF', size=11)

    # 참고 컬럼 헤더 회색 배경
    ref_start = len([c for c in core_cols + detail_cols if c in available_cols]) + 1
    for col_idx in range(ref_start, len(available_cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = PatternFill(start_color='6C757D', end_color='6C757D', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=10)


def create_hook_sheet(ws, hook_type_df, hook_strict_df=None):
    """훅 개선 효과 시트 생성 (가독성 개선)"""
    ws.title = "🔄 훅 개선 효과"

    if hook_type_df is None or len(hook_type_df) == 0:
        ws['A1'].value = "훅 비교 데이터가 없습니다."
        ws['A1'].font = Font(size=12, color='7F8C8D')
        return

    current_row = 1

    # Strict 매칭 결과 (있는 경우)
    if hook_strict_df is not None and len(hook_strict_df) > 0:
        # 섹션 헤더
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'].value = "🔍 정확 매칭 (동일 소재명)"
        ws[f'A{current_row}'].fill = SECTION_HEADER_FILL
        ws[f'A{current_row}'].font = SECTION_HEADER_FONT
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{current_row}'].fill = SECTION_HEADER_FILL
        ws.row_dimensions[current_row].height = 28
        current_row += 1

        strict_cols = ['소재유형', '소재명', 'CTR_변화', 'CVR_변화', 'CPA_변화', '훅판정']
        available_strict = [c for c in strict_cols if c in hook_strict_df.columns]

        if available_strict:
            # 헤더
            for c_idx, col in enumerate(available_strict, 1):
                ws.cell(row=current_row, column=c_idx, value=col)
            style_header(ws, current_row)
            header_row = current_row
            current_row += 1

            # 데이터 (변화율 포맷팅)
            for _, row_data in hook_strict_df[available_strict].iterrows():
                for c_idx, val in enumerate(row_data, 1):
                    col_name = available_strict[c_idx - 1]
                    cell = ws.cell(row=current_row, column=c_idx)

                    # 변화율 컬럼 포맷팅
                    if '변화' in col_name and col_name != '훅판정':
                        cell.value = format_change_rate(val)
                    else:
                        cell.value = val

                    cell.border = THIN_BORDER
                    ws.row_dimensions[current_row].height = 22
                current_row += 1

            # 변화율 색상 적용
            for c_idx, col_name in enumerate(available_strict, 1):
                if '변화' in col_name and col_name != '훅판정':
                    reverse = 'CPA' in col_name  # CPA는 하락이 좋음
                    apply_change_rate_colors(ws, c_idx, start_row=header_row + 1, reverse=reverse)

            # 훅판정 컬럼 스타일
            if '훅판정' in available_strict:
                verdict_col_idx = available_strict.index('훅판정') + 1
                apply_hook_verdict_styles(ws, verdict_col_idx, start_row=header_row + 1)

        current_row += 2

    # 소재유형별 집계 비교
    ws.merge_cells(f'A{current_row}:M{current_row}')
    ws[f'A{current_row}'].value = "📊 소재유형별 집계 비교"
    ws[f'A{current_row}'].fill = SECTION_HEADER_FILL
    ws[f'A{current_row}'].font = SECTION_HEADER_FONT
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws[f'{col}{current_row}'].fill = SECTION_HEADER_FILL
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    type_cols = ['소재유형', '신규_소재수', '신규_CTR', '신규_CVR', '신규_CPA',
                 '재가공_소재수', '재가공_CTR', '재가공_CVR', '재가공_CPA',
                 'CTR_변화율', 'CVR_변화율', 'CPA_변화율', '훅판정']
    available_type = [c for c in type_cols if c in hook_type_df.columns]

    # 헤더
    for c_idx, col in enumerate(available_type, 1):
        cell = ws.cell(row=current_row, column=c_idx, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    ws.row_dimensions[current_row].height = 25
    header_row_type = current_row
    current_row += 1

    # 데이터
    for _, row_data in hook_type_df[available_type].iterrows():
        for c_idx, val in enumerate(row_data, 1):
            col_name = available_type[c_idx - 1]
            cell = ws.cell(row=current_row, column=c_idx)

            # 변화율 컬럼 포맷팅
            if '변화율' in col_name:
                cell.value = format_change_rate(val)
            elif isinstance(val, float):
                if 'CPA' in col_name:
                    cell.value = val
                    cell.number_format = '#,##0'
                else:
                    cell.value = val
                    cell.number_format = '0.00'
            else:
                cell.value = val

            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right' if isinstance(val, (int, float)) else 'left',
                                        vertical='center')
        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # 변화율 색상 적용
    for c_idx, col_name in enumerate(available_type, 1):
        if '변화율' in col_name:
            reverse = 'CPA' in col_name
            apply_change_rate_colors(ws, c_idx, start_row=header_row_type + 1, reverse=reverse)

    # 훅판정 스타일
    if '훅판정' in available_type:
        verdict_col_idx = available_type.index('훅판정') + 1
        apply_hook_verdict_styles(ws, verdict_col_idx, start_row=header_row_type + 1)

    auto_column_width(ws)


def create_branch_sheet(ws, df_valid):
    """지점 컨텍스트 시트 생성 (순위 및 평가 추가)"""
    ws.title = "🏢 지점 컨텍스트"

    # 지점별 집계
    branch_summary = df_valid.groupby('지점').agg(
        총비용=('cost', 'sum'),
        총전환=('conversions', 'sum'),
        총클릭=('clicks', 'sum'),
        총노출=('impressions', 'sum'),
        소재수=('ad_name', 'nunique'),
        집행일수=('date', 'nunique'),
    ).reset_index()

    branch_summary['CPA'] = (branch_summary['총비용'] / branch_summary['총전환'].replace(0, np.nan)).round(0)
    branch_summary['CTR'] = (branch_summary['총클릭'] / branch_summary['총노출'].replace(0, np.nan) * 100).round(2)
    branch_summary['CVR'] = (branch_summary['총전환'] / branch_summary['총클릭'].replace(0, np.nan) * 100).round(2)

    # CPA 순으로 정렬 (성과 좋은 순)
    branch_summary = branch_summary.sort_values('CPA', na_position='last').reset_index(drop=True)

    # 순위 및 평가 컬럼 추가
    def get_rank_icon(idx):
        if idx == 0:
            return '🥇'
        elif idx == 1:
            return '🥈'
        elif idx == 2:
            return '🥉'
        else:
            return str(idx + 1)

    def get_evaluation(cpa, median_cpa):
        if pd.isna(cpa) or cpa == '':
            return ''
        if cpa <= median_cpa * 0.8:
            return '✅ 최우수'
        elif cpa <= median_cpa:
            return '👍 우수'
        elif cpa <= median_cpa * 1.2:
            return '➡️ 보통'
        else:
            return '⚠️ 개선필요'

    # 중앙값 계산 (평가 기준)
    valid_cpa = branch_summary['CPA'].dropna()
    median_cpa = valid_cpa.median() if len(valid_cpa) > 0 else 0

    branch_summary['순위'] = [get_rank_icon(i) for i in range(len(branch_summary))]
    branch_summary['평가'] = branch_summary['CPA'].apply(lambda x: get_evaluation(x, median_cpa))

    # 컬럼 순서 재정렬
    col_order = ['순위', '지점', '총비용', '총전환', 'CPA', 'CTR', 'CVR', '소재수', '집행일수', '평가']
    available_cols = [c for c in col_order if c in branch_summary.columns]
    branch_summary = branch_summary[available_cols]

    # NaN/inf 값 정리
    branch_summary = branch_summary.fillna('')

    df_to_sheet(ws, branch_summary)

    # 평가 컬럼 스타일 적용
    if '평가' in available_cols:
        eval_col_idx = available_cols.index('평가') + 1
        apply_efficiency_styles(ws, eval_col_idx, start_row=2)

    # 순위 컬럼 중앙 정렬
    if '순위' in available_cols:
        rank_col_idx = available_cols.index('순위') + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[rank_col_idx - 1].alignment = Alignment(horizontal='center', vertical='center')


def create_age_sheet(ws, age_df):
    """나이대 분석 시트 생성 (효율 판정 시각화)"""
    ws.title = "👥 나이대 분석"

    if len(age_df) == 0:
        ws['A1'].value = "나이대 데이터가 없습니다."
        ws['A1'].font = Font(size=12, color='7F8C8D')
        return

    # 출력 컬럼
    columns = ['age_group', '총비용', '총전환', '비용비중', '전환비중',
               'CPA', 'CTR', '예산효율점수', '효율판정', '신뢰도주의']
    available_cols = [c for c in columns if c in age_df.columns]
    df_output = age_df[available_cols].copy()

    # 컬럼명 한글화
    rename_map = {'age_group': '나이대'}
    df_output = df_output.rename(columns=rename_map)

    # 효율판정 아이콘 추가
    if '효율판정' in df_output.columns:
        def add_efficiency_icon(val):
            if pd.isna(val) or val == '':
                return ''
            val_str = str(val)
            if '우수' in val_str and '✅' not in val_str:
                return '✅ ' + val_str
            elif '양호' in val_str and '➡️' not in val_str:
                return '➡️ ' + val_str
            elif '낮음' in val_str and '⬇️' not in val_str:
                return '⬇️ ' + val_str
            elif '비효율' in val_str and '🔴' not in val_str:
                return '🔴 ' + val_str
            return val_str
        df_output['효율판정'] = df_output['효율판정'].apply(add_efficiency_icon)

    # 효율점수 기준 정렬 (높은 순)
    if '예산효율점수' in df_output.columns:
        df_output = df_output.sort_values('예산효율점수', ascending=False)

    df_to_sheet(ws, df_output)

    # 효율판정 컬럼 스타일
    if '효율판정' in df_output.columns:
        col_names = list(df_output.columns)
        eval_col_idx = col_names.index('효율판정') + 1
        apply_efficiency_styles(ws, eval_col_idx, start_row=2)

    # 비중 컬럼 시각화 (막대 효과)
    # 비용비중과 전환비중에 데이터 막대 효과 적용
    col_names = list(df_output.columns)
    if '비용비중' in col_names and '전환비중' in col_names:
        cost_idx = col_names.index('비용비중') + 1
        conv_idx = col_names.index('전환비중') + 1

        for row_num in range(2, ws.max_row + 1):
            # 비용비중: 파란색 배경 (비중에 따라 진하기 조절)
            cost_cell = ws.cell(row=row_num, column=cost_idx)
            if cost_cell.value and isinstance(cost_cell.value, (int, float)):
                intensity = min(int(cost_cell.value * 2), 100)
                if intensity > 0:
                    # 연한 파랑 배경
                    cost_cell.fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')

            # 전환비중: 초록색 배경
            conv_cell = ws.cell(row=row_num, column=conv_idx)
            if conv_cell.value and isinstance(conv_cell.value, (int, float)):
                intensity = min(int(conv_cell.value * 2), 100)
                if intensity > 0:
                    conv_cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')


def create_daily_trend_sheet(ws, df_valid):
    """일별 트렌드 시트 생성 (변화율 시각화)"""
    ws.title = "📅 일별 트렌드"

    # 일별 집계
    daily = df_valid.groupby('date').agg(
        총비용=('cost', 'sum'),
        총전환=('conversions', 'sum'),
        총클릭=('clicks', 'sum'),
        총노출=('impressions', 'sum'),
    ).reset_index().sort_values('date')

    daily['CPA'] = (daily['총비용'] / daily['총전환'].replace(0, np.nan)).round(0)
    daily['CTR'] = (daily['총클릭'] / daily['총노출'].replace(0, np.nan) * 100).round(2)
    daily['CVR'] = (daily['총전환'] / daily['총클릭'].replace(0, np.nan) * 100).round(2)

    # 전일 대비 변화율
    daily['CPA_전일대비'] = daily['CPA'].pct_change().mul(100).round(1)
    daily['전환_전일대비'] = daily['총전환'].pct_change().mul(100).round(1)

    # 변화율 포맷팅
    daily['CPA_변화'] = daily['CPA_전일대비'].apply(format_change_rate)
    daily['전환_변화'] = daily['전환_전일대비'].apply(format_change_rate)

    # 컬럼 순서 정리 (변화율 원본 제거)
    daily = daily.drop(columns=['CPA_전일대비', '전환_전일대비'], errors='ignore')

    # NaN/inf 값 정리
    daily = daily.fillna('')

    # 날짜 포맷 변환 (MM-DD 형식으로 간결하게)
    daily['날짜'] = pd.to_datetime(daily['date']).dt.strftime('%m-%d')
    daily = daily.drop(columns=['date'])

    # 컬럼 순서
    col_order = ['날짜', '총비용', '총전환', '총클릭', 'CPA', 'CPA_변화', 'CTR', 'CVR', '전환_변화']
    available_cols = [c for c in col_order if c in daily.columns]
    daily = daily[available_cols]

    df_to_sheet(ws, daily)

    # 변화율 컬럼 스타일 적용
    col_names = list(daily.columns)
    if 'CPA_변화' in col_names:
        cpa_change_idx = col_names.index('CPA_변화') + 1
        apply_change_rate_colors(ws, cpa_change_idx, start_row=2, reverse=True)  # CPA는 하락이 좋음

    if '전환_변화' in col_names:
        conv_change_idx = col_names.index('전환_변화') + 1
        apply_change_rate_colors(ws, conv_change_idx, start_row=2, reverse=False)  # 전환은 상승이 좋음


def create_off_sheet(ws, off_df):
    """OFF 소재 시트 생성 (마지막 성과 보존)"""
    ws.title = "⏸ OFF 소재"

    if off_df is None or len(off_df) == 0:
        # 안내 메시지
        ws.merge_cells('A1:F1')
        ws['A1'].value = "⏸ OFF 소재 (집행 종료)"
        ws['A1'].font = Font(bold=True, size=14, color='6C757D')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30

        ws['A3'].value = "현재 OFF 상태인 소재가 없습니다."
        ws['A3'].font = Font(size=11, color='7F8C8D')
        return

    # 시트 설명 헤더
    ws.merge_cells('A1:L1')
    ws['A1'].value = "⏸ OFF 소재 (집행 종료) - 마지막 성과 기록"
    ws['A1'].font = Font(bold=True, size=14, color='6C757D')
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 30

    # 출력 컬럼
    columns = ['소재유형', '소재명', 'CPA', 'CTR', 'CVR',
               '총비용', '총전환', '마지막집행일', '집행일수', '집행지점목록']
    available_cols = [c for c in columns if c in off_df.columns]
    df_output = off_df[available_cols].copy()

    # 집행지점목록을 문자열로 변환
    if '집행지점목록' in df_output.columns:
        df_output['집행지점목록'] = df_output['집행지점목록'].apply(
            lambda x: ', '.join(x) if isinstance(x, list) else str(x)
        )

    # 상태 컬럼 추가
    df_output.insert(0, '상태', '📴 OFF')

    # CPA 순 정렬 (성과 좋았던 순)
    if 'CPA' in df_output.columns:
        df_output = df_output.sort_values('CPA', na_position='last')

    # 데이터 작성 (3행부터)
    df = clean_dataframe_for_excel(df_output)

    # 헤더
    for c_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=3, column=c_idx, value=col)
        cell.fill = PatternFill(start_color='6C757D', end_color='6C757D', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 25

    # 데이터
    for r_idx, (_, row_data) in enumerate(df.iterrows(), 4):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            # OFF 상태 회색 배경
            cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

            # 숫자 포맷팅
            col_name = df.columns[c_idx - 1]
            if isinstance(val, float):
                if 'CPA' in col_name or '비용' in col_name:
                    cell.number_format = '#,##0'
                elif 'CTR' in col_name or 'CVR' in col_name:
                    cell.number_format = '0.00'

        ws.row_dimensions[r_idx].height = 20

    auto_column_width(ws)


def build_excel(output_dir: str, creative_df: pd.DataFrame, age_df: pd.DataFrame,
                df_valid: pd.DataFrame, off_df: pd.DataFrame = None,
                hook_type_df: pd.DataFrame = None, hook_strict_df: pd.DataFrame = None,
                target_cpa_map: dict = None):
    """
    Excel 리포트 생성 (7개 시트 + 11개 차트)

    v3.2.0: target_cpa_map 지원 (지점별 목표 CPA)
            신규 차트: 레이더, 피로도, CPA 목표선
    """
    if not OPENPYXL_AVAILABLE:
        print("[ERROR] openpyxl not installed. pip install openpyxl")
        return None

    wb = Workbook()

    # Prepare branch summary for charts
    branch_summary = df_valid.groupby('지점').agg(
        총비용=('cost', 'sum'),
        총전환=('conversions', 'sum'),
        총클릭=('clicks', 'sum'),
        총노출=('impressions', 'sum'),
    ).reset_index()
    branch_summary['CPA'] = (branch_summary['총비용'] / branch_summary['총전환'].replace(0, np.nan)).round(0)

    # Prepare daily summary for charts
    daily_summary = df_valid.groupby('date').agg(
        cost=('cost', 'sum'),
        conversions=('conversions', 'sum'),
        clicks=('clicks', 'sum'),
        impressions=('impressions', 'sum'),
    ).reset_index().sort_values('date')
    daily_summary['CTR'] = (daily_summary['clicks'] / daily_summary['impressions'].replace(0, np.nan) * 100).round(2)

    # Prepare pivot tables for heatmaps
    pivot_ctr = None
    pivot_cvr = None
    if '소재유형' in df_valid.columns and 'age_group' in df_valid.columns:
        try:
            pivot_ctr = df_valid.groupby(['소재유형', 'age_group']).agg(
                CTR=('CTR_calc', 'mean')
            ).reset_index().pivot(index='소재유형', columns='age_group', values='CTR').round(2)

            pivot_cvr = df_valid.groupby(['소재유형', 'age_group']).agg(
                CVR=('CVR_calc', 'mean')
            ).reset_index().pivot(index='소재유형', columns='age_group', values='CVR').round(2)
        except Exception as e:
            print(f"[WARNING] Pivot table creation failed: {e}")

    # 1. 요약 대시보드
    ws_summary = wb.active
    create_summary_sheet(ws_summary, creative_df, age_df, df_valid, off_df)

    # Add charts to summary sheet
    if CHARTS_AVAILABLE:
        try:
            add_tier_donut(ws_summary, creative_df, anchor_cell="K2")
            # v3.2.0: target_cpa_map 전달
            add_branch_cpa_bar(ws_summary, branch_summary, target_cpa_map=target_cpa_map, anchor_cell="K18")
            print("[Charts] Summary charts added")
        except Exception as e:
            print(f"[WARNING] Summary charts failed: {e}")

        # v3.2.0: 레이더 차트 추가
        try:
            add_type_radar_chart(ws_summary, creative_df, anchor_cell="K35")
            print("[Charts] Radar chart added")
        except Exception as e:
            print(f"[WARNING] Radar chart failed: {e}")

    # 2. 소재 TIER 분석
    ws_tier = wb.create_sheet()
    create_tier_sheet(ws_tier, creative_df)

    # Add bubble chart
    if CHARTS_AVAILABLE:
        try:
            anchor_row = ws_tier.max_row + 3
            add_creative_bubble_chart(ws_tier, creative_df, anchor_cell=f"A{anchor_row}")
            print("[Charts] Bubble chart added")
        except Exception as e:
            print(f"[WARNING] Bubble chart failed: {e}")

    # 3. 훅 개선 효과
    ws_hook = wb.create_sheet()
    create_hook_sheet(ws_hook, hook_type_df, hook_strict_df)

    # Add hook comparison chart
    if CHARTS_AVAILABLE and hook_type_df is not None and len(hook_type_df) > 0:
        try:
            anchor_row = ws_hook.max_row + 3
            add_hook_comparison_chart(ws_hook, hook_type_df, anchor_cell=f"A{anchor_row}")
            print("[Charts] Hook comparison chart added")
        except Exception as e:
            print(f"[WARNING] Hook chart failed: {e}")

    # 4. 지점 컨텍스트
    ws_branch = wb.create_sheet()
    create_branch_sheet(ws_branch, df_valid)

    # 5. 나이대 분석
    ws_age = wb.create_sheet()
    create_age_sheet(ws_age, age_df)

    # Add age efficiency chart and heatmaps
    if CHARTS_AVAILABLE:
        try:
            add_age_efficiency_chart(ws_age, age_df, anchor_cell="I3")
            print("[Charts] Age efficiency chart added")

            if pivot_ctr is not None and len(pivot_ctr) > 0:
                anchor_row = ws_age.max_row + 3
                add_age_type_heatmap(ws_age, pivot_ctr, 'CTR', anchor_cell=f"A{anchor_row}")
                print("[Charts] CTR heatmap added")

            if pivot_cvr is not None and len(pivot_cvr) > 0:
                anchor_row = ws_age.max_row + 20
                add_age_type_heatmap(ws_age, pivot_cvr, 'CVR', anchor_cell=f"A{anchor_row}")
                print("[Charts] CVR heatmap added")
        except Exception as e:
            print(f"[WARNING] Age charts failed: {e}")

    # 6. 일별 트렌드
    ws_daily = wb.create_sheet()
    create_daily_trend_sheet(ws_daily, df_valid)

    # Add daily charts
    if CHARTS_AVAILABLE:
        try:
            anchor_row = ws_daily.max_row + 3
            add_daily_combo_chart(ws_daily, daily_summary, anchor_cell=f"A{anchor_row}")
            print("[Charts] Daily combo chart added")

            anchor_row = ws_daily.max_row + 25
            add_daily_ctr_line(ws_daily, daily_summary, anchor_cell=f"A{anchor_row}")
            print("[Charts] Daily CTR chart added")
        except Exception as e:
            print(f"[WARNING] Daily charts failed: {e}")

        # v3.2.0: 피로도 차트 추가
        try:
            anchor_row = ws_daily.max_row + 20
            add_fatigue_line_chart(ws_daily, df_valid, creative_df, anchor_cell=f"A{anchor_row}")
            print("[Charts] Fatigue line chart added")
        except Exception as e:
            print(f"[WARNING] Fatigue chart failed: {e}")

        # v3.2.0: CPA 목표선 차트 추가
        try:
            # 평균 목표 CPA 계산 (있는 경우)
            avg_target_cpa = None
            if target_cpa_map and isinstance(target_cpa_map, dict):
                avg_target_cpa = np.mean(list(target_cpa_map.values()))
            elif target_cpa_map and isinstance(target_cpa_map, (int, float)):
                avg_target_cpa = target_cpa_map

            anchor_row = ws_daily.max_row + 30
            add_daily_cpa_trend_with_target(ws_daily, daily_summary, target_cpa=avg_target_cpa, anchor_cell=f"A{anchor_row}")
            print("[Charts] Daily CPA trend chart added")
        except Exception as e:
            print(f"[WARNING] Daily CPA chart failed: {e}")

    # 7. OFF 소재
    ws_off = wb.create_sheet()
    create_off_sheet(ws_off, off_df)

    # 저장
    os.makedirs(output_dir, exist_ok=True)
    today = datetime.now().strftime('%Y%m%d')
    output_path = os.path.join(output_dir, f"tiktok_analysis_{today}.xlsx")

    # Generate analysis_validation.json
    generate_validation_json(output_dir, df_valid, creative_df, age_df)

    wb.save(output_path)
    print(f"[OK] Excel report generated -> {output_path}")

    return output_path


def generate_validation_json(output_dir: str, df_valid: pd.DataFrame,
                            creative_df: pd.DataFrame, age_df: pd.DataFrame):
    """
    Generate analysis_validation.json for data integrity check
    """
    import json

    # Calculate totals
    raw_total_conv = df_valid['conversions'].sum()
    raw_total_cost = df_valid['cost'].sum()

    # Creative-level totals
    creative_conv = creative_df['총전환'].sum() if '총전환' in creative_df.columns else 0

    # Daily totals (from df_valid)
    daily_conv = df_valid.groupby('date')['conversions'].sum().sum()

    # Branch totals
    branch_conv = df_valid.groupby('지점')['conversions'].sum().sum() if '지점' in df_valid.columns else 0

    # Check matches
    all_match = (
        abs(raw_total_conv - creative_conv) <= 1 and
        abs(raw_total_conv - daily_conv) <= 1 and
        abs(raw_total_conv - branch_conv) <= 1
    )

    validation = {
        "generated_at": datetime.now().strftime('%Y-%m-%d %H:%M'),
        "validation": {
            "raw_total_conv": int(raw_total_conv),
            "raw_total_cost": float(raw_total_cost),
            "creative_conv_sum": int(creative_conv),
            "daily_conv_sum": int(daily_conv),
            "branch_conv_sum": int(branch_conv),
            "all_match": bool(all_match),  # Convert numpy bool to Python bool
            "analysis_total_conv": int(creative_conv)
        }
    }

    # Save
    os.makedirs(output_dir, exist_ok=True)
    json_path = os.path.join(output_dir, "analysis_validation.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(validation, f, ensure_ascii=False, indent=2)

    print(f"[OK] Validation JSON -> {json_path}")

    if not all_match:
        print(f"[WARNING] Data mismatch detected!")
        print(f"  Raw: {raw_total_conv}, Creative: {creative_conv}, Daily: {daily_conv}, Branch: {branch_conv}")

    return validation


def main(data_dir: str = "output", output_dir: str = None):
    """
    메인 실행 함수
    """
    if output_dir is None:
        today = datetime.now().strftime('%Y%m%d')
        output_dir = os.path.join(data_dir, today)

    # 데이터 로드
    creative_df = pd.read_parquet(os.path.join(data_dir, "creative_tier.parquet"))
    print(f"[Load] Creative TIER: {len(creative_df)}")

    try:
        age_df = pd.read_parquet(os.path.join(data_dir, "age_analysis.parquet"))
        print(f"[Load] Age analysis: {len(age_df)}")
    except FileNotFoundError:
        age_df = pd.DataFrame()

    try:
        off_df = pd.read_parquet(os.path.join(data_dir, "creative_off.parquet"))
        print(f"[Load] OFF creatives: {len(off_df)}")
    except FileNotFoundError:
        off_df = pd.DataFrame()

    try:
        df_valid = pd.read_parquet(os.path.join(data_dir, "parsed.parquet"))
        df_valid = df_valid[df_valid['parse_status'] == 'OK']
        print(f"[Load] Parsed data: {len(df_valid)} rows")
    except FileNotFoundError:
        print("[ERROR] parsed.parquet not found")
        return None

    # 훅 비교 데이터 로드
    try:
        hook_type_df = pd.read_parquet(os.path.join(data_dir, "hook_type_comparison.parquet"))
        print(f"[Load] Hook type comparison: {len(hook_type_df)}")
    except FileNotFoundError:
        hook_type_df = None

    try:
        hook_strict_df = pd.read_parquet(os.path.join(data_dir, "hook_strict_pairs.parquet"))
        print(f"[Load] Hook strict pairs: {len(hook_strict_df)}")
    except FileNotFoundError:
        hook_strict_df = None

    return build_excel(output_dir, creative_df, age_df, df_valid, off_df, hook_type_df, hook_strict_df)


if __name__ == "__main__":
    import sys
    data_directory = sys.argv[1] if len(sys.argv) > 1 else "output"
    output_directory = sys.argv[2] if len(sys.argv) > 2 else None
    main(data_directory, output_directory)
